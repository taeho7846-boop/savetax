"""
CMS 신청서 엑셀 → PDF 생성
Usage: python generate_cms_form.py <template_path> <output_pdf_path>
       <first_month> <depositor> <resident6> <bank_name> <biz_number>
       <bank_account> <phone> <stamp_name> <client_type>
"""

import sys
import os
import tempfile
import shutil
import subprocess
from pathlib import Path


def _get_font(font_size, bold=True):
    from PIL import ImageFont
    if bold:
        font_paths = [
            "/usr/share/fonts/truetype/nanum/NanumGothicExtraBold.ttf",
            "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf",
            "/usr/share/fonts/nanum/NanumGothicExtraBold.ttf",
            "/usr/share/fonts/nanum/NanumGothicBold.ttf",
            "C:/Windows/Fonts/malgunbd.ttf",
        ]
    else:
        font_paths = [
            "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf",
            "/usr/share/fonts/nanum/NanumGothicBold.ttf",
            "C:/Windows/Fonts/malgunbd.ttf",
            "C:/Windows/Fonts/malgun.ttf",
        ]
    for fp in font_paths:
        try:
            return ImageFont.truetype(fp, font_size)
        except Exception:
            pass
    return ImageFont.load_default()


def create_personal_stamp(name, size_px=300):
    """개인 도장: 원형 + 이름 세로 배치"""
    from PIL import Image, ImageDraw
    import io

    img = Image.new("RGBA", (size_px, size_px), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)

    margin = size_px // 15
    line_w = round(15 * size_px / 300)
    draw.ellipse(
        [margin, margin, size_px - margin, size_px - margin],
        outline=(180, 0, 0),
        width=line_w,
    )

    n = len(name)
    inner_h = size_px - margin * 5
    font_size = max(14, int(inner_h / (n + 0.3)))
    font = _get_font(font_size)

    char_sizes = []
    for char in name:
        bbox = draw.textbbox((0, 0), char, font=font)
        char_sizes.append((bbox[2] - bbox[0], bbox[3] - bbox[1]))

    total_h = sum(h for _, h in char_sizes)
    y = (size_px - total_h) // 2
    for char, (cw, ch) in zip(name, char_sizes):
        x = (size_px - cw) // 2
        draw.text((x, y), char, fill=(180, 0, 0), font=font)
        y += ch

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def create_corporate_stamp(name, size_px=300):
    """법인 도장: 외곽에 회사명 원형 배치 + 중앙에 대표이사"""
    from PIL import Image, ImageDraw
    import io
    import math

    img = Image.new("RGBA", (size_px, size_px), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)

    margin = size_px // 15
    line_w = round(15 * size_px / 300)
    center = size_px // 2
    radius = center - margin
    color = (180, 0, 0)

    # 외곽 원 (이중 원)
    outer_lw = max(3, line_w // 2)
    draw.ellipse(
        [margin, margin, size_px - margin, size_px - margin],
        outline=color, width=outer_lw,
    )
    inner_margin = margin + outer_lw + 3
    draw.ellipse(
        [inner_margin, inner_margin, size_px - inner_margin, size_px - inner_margin],
        outline=color, width=max(2, outer_lw // 2),
    )

    # 중앙 가로선 (대표이사 영역 구분)
    center_box_h = size_px // 4
    top_line_y = center - center_box_h // 2
    bot_line_y = center + center_box_h // 2
    line_x_margin = inner_margin + 8
    draw.line([(line_x_margin, top_line_y), (size_px - line_x_margin, top_line_y)], fill=color, width=max(2, outer_lw // 2))
    draw.line([(line_x_margin, bot_line_y), (size_px - line_x_margin, bot_line_y)], fill=color, width=max(2, outer_lw // 2))

    # 중앙 "대표이사" 가로
    center_text = "대표이사"
    center_font_size = center_box_h * 3 // 5
    center_font = _get_font(center_font_size)
    bbox_ct = draw.textbbox((0, 0), center_text, font=center_font)
    ct_w = bbox_ct[2] - bbox_ct[0]
    ct_h = bbox_ct[3] - bbox_ct[1]
    draw.text(
        ((size_px - ct_w) // 2, center - ct_h // 2),
        center_text, fill=color, font=center_font,
    )

    # 외곽 회사명 원형 배치
    chars = list(name)
    n = len(chars)
    if n > 0:
        outer_font_size = max(20, int((size_px * 1.1) / (n + 0.5)))
        outer_font = _get_font(outer_font_size)
        text_radius = radius - line_w - margin

        # 12시 방향부터 시계방향
        start_angle = -math.pi / 2
        angle_step = (2 * math.pi) / n

        for i, char in enumerate(chars):
            angle = start_angle + angle_step * i
            cx = center + text_radius * math.cos(angle)
            cy = center + text_radius * math.sin(angle)

            bbox = draw.textbbox((0, 0), char, font=outer_font)
            cw = bbox[2] - bbox[0]
            ch_h = bbox[3] - bbox[1]

            # 글자 개별 이미지 생성 + 회전
            pad = 6
            char_img = Image.new("RGBA", (cw + pad * 2, ch_h + pad * 2), (0, 0, 0, 0))
            char_draw = ImageDraw.Draw(char_img)
            char_draw.text((pad, pad), char, fill=color, font=outer_font)

            rot_deg = -math.degrees(angle) - 90
            char_img = char_img.rotate(rot_deg, expand=True, resample=Image.BICUBIC)

            paste_x = int(cx - char_img.width / 2)
            paste_y = int(cy - char_img.height / 2)
            img.paste(char_img, (paste_x, paste_y), char_img)

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def create_stamp_png(name, size_px=300, is_corporate=False):
    if is_corporate:
        return create_corporate_stamp(name, size_px)
    return create_personal_stamp(name, size_px)


def main():
    import io as _io
    sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = _io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

    if len(sys.argv) < 12:
        print("ERROR: 인수 부족", file=sys.stderr)
        sys.exit(1)

    template_path = sys.argv[1]
    output_pdf    = sys.argv[2]
    first_month   = sys.argv[3]   # 202604
    depositor     = sys.argv[4]   # 예금주명
    resident6     = sys.argv[5]   # 주민번호앞6 (개인) or ""
    bank_name     = sys.argv[6]   # 은행명
    biz_number    = sys.argv[7]   # 사업자번호 (법인) or ""
    bank_account  = sys.argv[8]   # 계좌번호
    phone         = sys.argv[9]   # 연락처
    stamp_name    = sys.argv[10]  # 도장 이름
    client_type   = sys.argv[11]  # individual or corporate

    if not os.path.isfile(template_path):
        print(f"ERROR: 템플릿 파일 없음: {template_path}", file=sys.stderr)
        sys.exit(1)

    # 1. 도장 생성
    try:
        is_corp = client_type == "corporate"
        stamp_data = create_stamp_png(stamp_name, size_px=300, is_corporate=is_corp)
        print(f"도장 생성: {stamp_name} ({'법인' if is_corp else '개인'})")
    except Exception as e:
        print(f"ERROR: 도장 생성 실패: {e}", file=sys.stderr)
        sys.exit(1)

    # 2. 템플릿 복사 + openpyxl로 셀 채우기
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import cm_to_EMU
    import io

    ext = Path(template_path).suffix or ".xlsx"
    tmp_fd, tmp_xlsx = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    shutil.copy2(template_path, tmp_xlsx)

    try:
        wb = load_workbook(tmp_xlsx)
        ws = wb.active

        ws["B9"] = first_month
        ws["B14"] = depositor
        ws["D14"] = resident6 if client_type == "individual" else ""
        ws["B15"] = bank_name
        ws["D15"] = biz_number if client_type == "corporate" else ""
        ws["B16"] = bank_account
        ws["D17"] = phone
        print("셀 입력 완료")

        # 도장 삽입 D28
        STAMP_CM = 2.0 if client_type == "individual" else 2.5
        stamp_emu = cm_to_EMU(STAMP_CM)
        stamp_img = XlImage(io.BytesIO(stamp_data))
        stamp_img.width = STAMP_CM / 2.54 * 96
        stamp_img.height = STAMP_CM / 2.54 * 96
        col_off = 2.0 if client_type == "individual" else 2.5
        marker = AnchorMarker(col=3, colOff=cm_to_EMU(col_off), row=27, rowOff=cm_to_EMU(0.0))
        anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(stamp_emu, stamp_emu))
        stamp_img.anchor = anchor
        ws.add_image(stamp_img)
        print("도장 삽입: D28")

        wb.save(tmp_xlsx)
    except Exception as e:
        print(f"ERROR: 엑셀 작성 실패: {e}", file=sys.stderr)
        os.unlink(tmp_xlsx)
        sys.exit(1)

    # 3. PDF 변환 (LibreOffice)
    try:
        output_dir = os.path.dirname(os.path.abspath(output_pdf))
        os.makedirs(output_dir, exist_ok=True)

        result = subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf", "--outdir", output_dir, tmp_xlsx],
            capture_output=True, text=True, timeout=60
        )
        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice 변환 실패: {result.stderr}")

        generated = os.path.join(output_dir, Path(tmp_xlsx).stem + ".pdf")
        if generated != os.path.abspath(output_pdf):
            shutil.move(generated, os.path.abspath(output_pdf))

        print(f"SUCCESS: {output_pdf}")
    except Exception as e:
        print(f"ERROR: PDF 변환 실패: {e}", file=sys.stderr)
        sys.exit(1)
    finally:
        try:
            os.unlink(tmp_xlsx)
        except Exception:
            pass


if __name__ == "__main__":
    main()
