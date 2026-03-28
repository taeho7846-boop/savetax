"""
홈택스수임신청서 엑셀 템플릿 → PDF 생성
Usage: python generate_commission_form.py <template_path> <output_pdf_path>
       <ceo_name> <resident_number> <client_name> <biz_number> <phone>

Linux(VPS): openpyxl + LibreOffice
Windows: openpyxl + win32com (Excel COM)
"""

import sys
import os
import tempfile
import shutil
import subprocess
from pathlib import Path


def create_stamp_png(name: str, size_px: int = 300) -> bytes:
    """대표자 이름 원형 도장 PNG 생성"""
    from PIL import Image, ImageDraw, ImageFont
    import io

    img = Image.new("RGBA", (size_px, size_px), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)

    margin = size_px // 15
    line_w = round(15 * size_px / 300)
    draw.ellipse(
        [margin, margin, size_px - margin, size_px - margin],
        outline=(180, 0, 0),
        width=line_w,
    )

    n = len(name)
    inner_h = size_px - margin * 5
    font_size = max(14, int(inner_h / (n + 0.3)))

    font_paths_linux = [
        "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf",
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
        "/usr/share/fonts/nanum/NanumGothicBold.ttf",
        "/usr/share/fonts/nanum/NanumGothic.ttf",
    ]
    font_paths_win = [
        "C:/Windows/Fonts/malgunbd.ttf",
        "C:/Windows/Fonts/malgun.ttf",
        "C:/Windows/Fonts/gulim.ttc",
        "C:/Windows/Fonts/batang.ttc",
    ]
    font_paths = font_paths_linux + font_paths_win

    font = None
    for fp in font_paths:
        try:
            font = ImageFont.truetype(fp, font_size)
            break
        except Exception:
            pass
    if font is None:
        font = ImageFont.load_default()

    char_sizes = []
    for char in name:
        bbox = draw.textbbox((0, 0), char, font=font)
        char_sizes.append((bbox[2] - bbox[0], bbox[3] - bbox[1]))

    total_h = sum(h for _, h in char_sizes)
    y = (size_px - total_h) // 2
    for char, (cw, ch) in zip(name, char_sizes):
        x = (size_px - cw) // 2
        draw.text((x, y), char, fill=(180, 0, 0), font=font)
        y += ch

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def convert_with_libreoffice(xlsx_path: str, output_pdf: str):
    """LibreOffice로 xlsx → pdf 변환 (Linux)"""
    output_dir = os.path.dirname(os.path.abspath(output_pdf))
    os.makedirs(output_dir, exist_ok=True)

    result = subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "pdf", "--outdir", output_dir, xlsx_path],
        capture_output=True, text=True, timeout=60
    )
    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice 변환 실패: {result.stderr}")

    # LibreOffice는 원본 파일명.pdf로 생성하므로 원하는 이름으로 이동
    generated = os.path.join(output_dir, Path(xlsx_path).stem + ".pdf")
    if generated != os.path.abspath(output_pdf):
        shutil.move(generated, os.path.abspath(output_pdf))


def convert_with_win32com(xlsx_path: str, output_pdf: str):
    """win32com(Excel COM)으로 xlsx → pdf 변환 (Windows)"""
    import win32com.client

    os.makedirs(os.path.dirname(os.path.abspath(output_pdf)), exist_ok=True)

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
        wb.ExportAsFixedFormat(0, os.path.abspath(output_pdf))
        wb.Close(False)
    finally:
        excel.Quit()


def fill_xlsx_with_openpyxl(template_path, tmp_xlsx, ceo_name, resident_number, client_name, biz_number, phone, stamp_data):
    """openpyxl로 엑셀 셀 채우기 + 도장 삽입"""
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import cm_to_EMU
    import io

    wb = load_workbook(template_path)
    ws = wb.active

    # 셀 채우기
    ws["B6"] = ceo_name
    ws["C6"] = resident_number
    ws["B8"] = client_name
    ws["C8"] = biz_number
    ws["D11"] = phone
    print("셀 입력 완료")

    # 도장 삽입 (EMU 단위: 1cm = 360000 EMU)
    STAMP_CM = 2.0
    stamp_emu = cm_to_EMU(STAMP_CM)

    # (col 0-indexed, col_offset_cm, row 0-indexed, row_offset_cm)
    stamp_positions = [
        (3, 1.5, 29, 0.0),   # D30 기준 오른쪽 1.5cm
        (3, 1.5, 34, 0.0),   # D35 기준 오른쪽 1.5cm
        (1, 1.5, 43, 0.0),   # B44 기준 오른쪽 1.5cm
    ]

    for col, col_off_cm, row, row_off_cm in stamp_positions:
        stamp_img = XlImage(io.BytesIO(stamp_data))
        stamp_img.width = STAMP_CM / 2.54 * 96
        stamp_img.height = STAMP_CM / 2.54 * 96
        marker = AnchorMarker(
            col=col,
            colOff=cm_to_EMU(col_off_cm),
            row=row,
            rowOff=cm_to_EMU(row_off_cm),
        )
        anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(stamp_emu, stamp_emu))
        stamp_img.anchor = anchor
        ws.add_image(stamp_img)
        print(f"도장 삽입: col={col} row={row} offset=({col_off_cm}cm, {row_off_cm}cm)")

    wb.save(tmp_xlsx)
    print(f"엑셀 저장: {tmp_xlsx}")


def main():
    import io as _io
    sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = _io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

    if len(sys.argv) < 8:
        print("ERROR: 인수 부족", file=sys.stderr)
        sys.exit(1)

    template_path   = sys.argv[1]
    output_pdf      = sys.argv[2]
    ceo_name        = sys.argv[3]
    resident_number = sys.argv[4]
    client_name     = sys.argv[5]
    biz_number      = sys.argv[6]
    phone           = sys.argv[7]

    if not os.path.isfile(template_path):
        print(f"ERROR: 템플릿 파일 없음: {template_path}", file=sys.stderr)
        sys.exit(1)

    # 1. 도장 이미지 생성
    try:
        stamp_data = create_stamp_png(ceo_name, size_px=300)
        print(f"도장 생성: {ceo_name}")
    except ImportError:
        print("ERROR: Pillow를 설치해주세요 (pip install Pillow)", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: 도장 생성 실패: {e}", file=sys.stderr)
        sys.exit(1)

    # 2. 템플릿 복사 + openpyxl로 셀/도장 채우기
    ext = Path(template_path).suffix or ".xlsx"
    tmp_fd, tmp_xlsx = tempfile.mkstemp(suffix=ext)
    os.close(tmp_fd)
    shutil.copy2(template_path, tmp_xlsx)

    try:
        fill_xlsx_with_openpyxl(template_path, tmp_xlsx, ceo_name, resident_number, client_name, biz_number, phone, stamp_data)
    except ImportError:
        print("ERROR: openpyxl을 설치해주세요 (pip install openpyxl)", file=sys.stderr)
        os.unlink(tmp_xlsx)
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: 엑셀 작성 실패: {e}", file=sys.stderr)
        os.unlink(tmp_xlsx)
        sys.exit(1)

    # 3. PDF 변환
    try:
        is_windows = sys.platform == "win32"
        if is_windows:
            try:
                convert_with_win32com(tmp_xlsx, output_pdf)
                print("PDF 변환 완료 (Excel COM)")
            except ImportError:
                convert_with_libreoffice(tmp_xlsx, output_pdf)
                print("PDF 변환 완료 (LibreOffice)")
        else:
            convert_with_libreoffice(tmp_xlsx, output_pdf)
            print("PDF 변환 완료 (LibreOffice)")

        print(f"SUCCESS: {output_pdf}")
    except Exception as e:
        print(f"ERROR: PDF 변환 실패: {e}", file=sys.stderr)
        sys.exit(1)
    finally:
        try:
            os.unlink(tmp_xlsx)
        except Exception:
            pass


if __name__ == "__main__":
    main()
