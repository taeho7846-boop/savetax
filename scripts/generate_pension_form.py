"""
국민연금 신고서 엑셀 → PDF 생성
Usage: python generate_pension_form.py <template_path> <output_pdf_path>
       <biz_number> <client_name> <address> <corporate_number>
       <ceo_name> <birthday> <stamp_name> <client_type>
"""

import sys
import os
import tempfile
import shutil
import subprocess
from pathlib import Path


def create_stamp_png(name, size_px=300):
    from PIL import Image, ImageDraw, ImageFont
    import io

    img = Image.new("RGBA", (size_px, size_px), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)
    margin = size_px // 15
    line_w = round(15 * size_px / 300)
    draw.ellipse([margin, margin, size_px - margin, size_px - margin], outline=(180, 0, 0), width=line_w)

    n = len(name)
    inner_h = size_px - margin * 5
    font_size = max(14, int(inner_h / (n + 0.3)))

    font_paths = [
        "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf",
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
        "/usr/share/fonts/nanum/NanumGothicBold.ttf",
        "C:/Windows/Fonts/malgunbd.ttf",
        "C:/Windows/Fonts/malgun.ttf",
    ]
    font = None
    for fp in font_paths:
        try:
            font = ImageFont.truetype(fp, font_size)
            break
        except Exception:
            pass
    if font is None:
        font = ImageFont.load_default()

    char_sizes = []
    for char in name:
        bbox = draw.textbbox((0, 0), char, font=font)
        char_sizes.append((bbox[2] - bbox[0], bbox[3] - bbox[1]))

    total_h = sum(h for _, h in char_sizes)
    y = (size_px - total_h) // 2
    for char, (cw, ch) in zip(name, char_sizes):
        x = (size_px - cw) // 2
        draw.text((x, y), char, fill=(180, 0, 0), font=font)
        y += ch

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def format_biz_f11(biz):
    biz = biz.replace("-", "").replace(" ", "")
    if len(biz) >= 10:
        return f"{biz[:3]}-{biz[3:5]}-{biz[5:10]}-0"
    return biz + "-0"

def format_biz_f13(biz):
    biz = biz.replace("-", "").replace(" ", "")
    if len(biz) >= 10:
        return f"{biz[:3]}-{biz[3:5]}-{biz[5:10]}"
    return biz

def format_corp_num(num):
    num = num.replace("-", "").replace(" ", "")
    if len(num) >= 13:
        return f"{num[:6]}-{num[6:13]}"
    return num


def fill_xlsx(template_path, tmp_xlsx, biz_number, client_name, address,
              corporate_number, ceo_name, birthday, stamp_data, client_type):
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import cm_to_EMU
    import io

    wb = load_workbook(template_path)
    ws = wb.active

    is_corp = client_type == "corporate"

    ws["F11"] = format_biz_f11(biz_number)
    ws["N11"] = client_name
    ws["F12"] = address
    ws["F13"] = format_biz_f13(biz_number)
    ws["N13"] = format_corp_num(corporate_number) if is_corp and corporate_number else ""
    ws["F15"] = ceo_name
    ws["M15"] = birthday
    print("셀 입력 완료")

    # 도장 삽입 Q31
    STAMP_CM = 2.0
    stamp_emu = cm_to_EMU(STAMP_CM)
    stamp_img = XlImage(io.BytesIO(stamp_data))
    stamp_img.width = STAMP_CM / 2.54 * 96
    stamp_img.height = STAMP_CM / 2.54 * 96
    marker = AnchorMarker(col=16, colOff=cm_to_EMU(0.5), row=30, rowOff=cm_to_EMU(0.0))
    anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(stamp_emu, stamp_emu))
    stamp_img.anchor = anchor
    ws.add_image(stamp_img)
    print("도장 삽입: Q31")

    wb.save(tmp_xlsx)
    print(f"엑셀 저장: {tmp_xlsx}")


def main():
    import io as _io
    sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = _io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

    if len(sys.argv) < 11:
        print("ERROR: 인수 부족", file=sys.stderr)
        sys.exit(1)

    template_path    = sys.argv[1]
    output_pdf       = sys.argv[2]
    biz_number       = sys.argv[3]
    client_name      = sys.argv[4]
    address          = sys.argv[5]
    corporate_number = sys.argv[6]
    ceo_name         = sys.argv[7]
    birthday         = sys.argv[8]
    stamp_name       = sys.argv[9]
    client_type      = sys.argv[10]

    if not os.path.isfile(template_path):
        print(f"ERROR: 템플릿 파일 없음: {template_path}", file=sys.stderr)
        sys.exit(1)

    # 1. 도장 생성
    try:
        stamp_data = create_stamp_png(stamp_name, size_px=300)
        print(f"도장 생성: {stamp_name}")
    except Exception as e:
        print(f"ERROR: 도장 생성 실패: {e}", file=sys.stderr)
        sys.exit(1)

    # 2. 템플릿 복사 + openpyxl로 셀/도장 채우기
    ext = Path(template_path).suffix or ".xlsx"
    tmp_fd, tmp_xlsx = tempfile.mkstemp(suffix=ext)
    os.close(tmp_fd)
    shutil.copy2(template_path, tmp_xlsx)

    try:
        fill_xlsx(template_path, tmp_xlsx, biz_number, client_name, address,
                  corporate_number, ceo_name, birthday, stamp_data, client_type)
    except Exception as e:
        print(f"ERROR: 엑셀 작성 실패: {e}", file=sys.stderr)
        os.unlink(tmp_xlsx)
        sys.exit(1)

    # 3. PDF 변환 (LibreOffice)
    try:
        output_dir = os.path.dirname(os.path.abspath(output_pdf))
        os.makedirs(output_dir, exist_ok=True)
        result = subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf", "--outdir", output_dir, tmp_xlsx],
            capture_output=True, text=True, timeout=60
        )
        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice 변환 실패: {result.stderr}")
        generated = os.path.join(output_dir, Path(tmp_xlsx).stem + ".pdf")
        if generated != os.path.abspath(output_pdf):
            shutil.move(generated, os.path.abspath(output_pdf))
        print(f"SUCCESS: {output_pdf}")
    except Exception as e:
        print(f"ERROR: PDF 변환 실패: {e}", file=sys.stderr)
        sys.exit(1)
    finally:
        try:
            os.unlink(tmp_xlsx)
        except Exception:
            pass


if __name__ == "__main__":
    main()
