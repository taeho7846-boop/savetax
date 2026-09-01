# 도움컴퍼니 원천세 월간 변환 도구
# 사용법:
#   python scripts/doum/doum-convert.py "<정산원본.xlsx>" <YYYY.MM> ["<일용직사원등록.xlsx>"]
# 예시:
#   python scripts/doum/doum-convert.py "G:\...\26년09월 원천세_일용직_도움컴퍼니.xlsx" 2026.09 "C:\Users\aaron\Downloads\일용직사원등록_20261001.xlsx"
#
# 산출물 (정산원본과 같은 폴더에 저장):
#   1. 사업소득자료입력_도움컴퍼니_YYMM.xls   — 위하고 SmartA 업로드 서식
#   2. 일용직급여자료입력_도움컴퍼니_YYMM.xls — 위하고 SmartA 업로드 서식 (같은 사람+같은 일자 합산)
#   3. 주민번호검증_도움컴퍼니_YYMM.xlsx      — 주민번호 오류/의심 목록
#   4. 일용직_신규등록대상_YYMM.xlsx          — 위하고에 없는 일용직 (직접 등록 필요, 사원등록 파일 준 경우만)
import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
import xlrd
from xlutils.copy import copy as xlcopy
from datetime import datetime, date

TPL_DIR = os.path.dirname(os.path.abspath(__file__))
TPL_BIZ = os.path.join(TPL_DIR, "사업소득자료입력_template.xls")
TPL_DAILY = os.path.join(TPL_DIR, "일용직급여자료입력_template.xls")


def clean_resident(v):
    return str(v or "").replace("-", "").strip()


def fmt_date(v, fallback_ym):
    if isinstance(v, datetime):
        return f"{v.year}.{v.month:02d}.{v.day:02d}"
    s = str(v or "")[:10].replace("-", ".")
    return s if len(s) == 10 else fallback_ym + ".01"


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def floor10(x):
    return int(x // 10) * 10


def check_rrn(rrn, today):
    issues = []
    d = rrn.replace("-", "").strip()
    if not d.isdigit() or len(d) != 13:
        return [f"자릿수 오류 ({len(d)}자리)"]
    g = d[6]
    cent = {"1": 1900, "2": 1900, "3": 2000, "4": 2000, "5": 1900, "6": 1900, "7": 2000, "8": 2000, "9": 1800, "0": 1800}.get(g)
    if cent is None:
        return [f"성별자리 오류 (7번째={g})"]
    try:
        birth = date(cent + int(d[0:2]), int(d[2:4]), int(d[4:6]))
    except ValueError:
        return [f"생년월일 불가 ({d[0:6]})"]
    if birth > today:
        issues.append("생년월일이 미래 (성별자리 확인)")
    else:
        age = (today - birth).days // 365
        if age < 14:
            issues.append(f"나이 이상 (만 {age}세)")
        elif age > 90:
            issues.append(f"고령 확인 필요 (만 {age}세)")
    w = [2, 3, 4, 5, 6, 7, 8, 9, 2, 3, 4, 5]
    s = sum(int(d[i]) * w[i] for i in range(12))
    if (11 - s % 11) % 10 != int(d[12]):
        issues.append("체크섬 불일치(오타 의심)" if g in "1234" else "체크섬 불일치(외국인 신규발급이면 정상 가능)")
    return issues


def main():
    if len(sys.argv) < 3:
        print(__doc__ or "사용법: python doum-convert.py <정산원본.xlsx> <YYYY.MM> [일용직사원등록.xlsx]")
        sys.exit(1)
    src_path = sys.argv[1]
    ym = sys.argv[2]  # "2026.09"
    reg_path = sys.argv[3] if len(sys.argv) > 3 else None
    year, month = ym.split(".")
    yymm = year[2:] + month
    out_dir = os.path.dirname(os.path.abspath(src_path))
    today = date.today()

    wb = openpyxl.load_workbook(src_path, data_only=True)

    # ── 사업소득 (3.3원천세) ──
    biz = []
    for r in wb["3.3원천세"].iter_rows(min_row=3, values_only=True):
        name = str(r[2] or "").strip()
        if not name:
            continue
        amount = num(r[9])
        tax = num(r[11])
        local = num(r[13])
        # 원본이 수식 셀이라 값이 비어 읽히면 규칙대로 재계산 (3% 절사, 지방세는 그 10% 절사)
        if amount > 0 and tax == 0:
            tax = floor10(amount * 0.03)
        if tax > 0 and local == 0:
            local = floor10(tax * 0.1)
        biz.append({
            "date": fmt_date(r[16], ym), "name": name, "resident": clean_resident(r[5]),
            "amount": amount, "tax": tax, "local": local, "work": str(r[1] or "").strip(),
        })

    # ── 일용직 (일용직고용보험) — 같은 사람+같은 일자 합산 ──
    daily_raw = []
    for r in wb["일용직고용보험"].iter_rows(min_row=3, values_only=True):
        name = str(r[2] or "").strip()
        if not name:
            continue
        d = r[12]
        day = d.day if isinstance(d, datetime) else int(str(d)[8:10])
        d_amount = num(r[9])
        d_emp = num(r[11])
        # 고용보험(0.9% 10원 절사)이 수식 셀이라 비어 읽히면 재계산
        if d_amount > 0 and d_emp == 0:
            d_emp = floor10(d_amount * 0.009)
        daily_raw.append({
            "day": day, "name": name, "resident": clean_resident(r[5]),
            "rrn_disp": str(r[5] or "").strip(), "work": str(r[1] or "").strip(),
            "bank": str(r[3] or "").strip(), "account": str(r[4] or "").strip(),
            "amount": d_amount, "emp": d_emp,
        })
    merged = {}
    for p in daily_raw:
        key = (p["resident"], p["day"])
        if key in merged:
            merged[key]["amount"] += p["amount"]
            merged[key]["emp"] += p["emp"]
        else:
            merged[key] = dict(p)
    daily = sorted(merged.values(), key=lambda p: (p["day"], p["name"]))

    # ── 1) 사업소득 서식 ──
    rb = xlrd.open_workbook(TPL_BIZ, formatting_info=True)
    wbo = xlcopy(rb)
    ws = wbo.get_sheet(0)
    for i, p in enumerate(biz):
        r = 3 + i
        ws.write(r, 0, i + 1); ws.write(r, 1, ym); ws.write(r, 2, p["date"])
        ws.write(r, 3, p["name"]); ws.write(r, 4, p["resident"])
        ws.write(r, 7, "기타인적용역자"); ws.write(r, 8, p["date"])
        ws.write(r, 9, p["amount"]); ws.write(r, 10, 3.0)
        ws.write(r, 11, p["tax"]); ws.write(r, 12, p["local"]); ws.write(r, 13, 0.0)
    out_biz = os.path.join(out_dir, f"사업소득자료입력_도움컴퍼니_{yymm}.xls")
    wbo.save(out_biz)

    # ── 2) 일용직 서식 ──
    rb2 = xlrd.open_workbook(TPL_DAILY, formatting_info=True)
    wbo2 = xlcopy(rb2)
    ws2 = wbo2.get_sheet(0)
    for i, p in enumerate(daily):
        r = 4 + i
        ws2.write(r, 0, i + 1); ws2.write(r, 1, p["day"]); ws2.write(r, 2, 1)
        ws2.write(r, 3, p["name"]); ws2.write(r, 4, p["resident"])
        ws2.write(r, 7, 8.0); ws2.write(r, 9, p["amount"]); ws2.write(r, 12, p["emp"])
        ws2.write(r, 16, 0.0); ws2.write(r, 17, 0.0)
    out_daily = os.path.join(out_dir, f"일용직급여자료입력_도움컴퍼니_{yymm}.xls")
    wbo2.save(out_daily)

    # ── 3) 주민번호 검증 ──
    people = [("사업소득", p["work"], p["name"], p["resident"]) for p in biz] + \
             [("일용직", p["work"], p["name"], p["resident"]) for p in daily]
    issues = []
    for tag, work, name, rrn in people:
        for msg in check_rrn(rrn, today):
            issues.append((tag, work, name, rrn, msg))
    by_rrn = {}
    for tag, work, name, rrn in people:
        by_rrn.setdefault(rrn, set()).add(name)
    for rrn, names in by_rrn.items():
        if len(names) > 1:
            issues.append(("공통", "", "/".join(sorted(names)), rrn, "같은 주민번호에 다른 이름 (오타 의심)"))
    vout = openpyxl.Workbook()
    vws = vout.active
    vws.title = "검증결과"
    vws.append(["구분", "근무처", "이름", "주민번호", "문제"])
    for c in vws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
    seen = set()
    for row in issues:
        if row in seen:
            continue
        seen.add(row)
        vws.append(list(row))
    for col, w in zip("ABCDE", [8, 18, 14, 18, 45]):
        vws.column_dimensions[col].width = w
    out_verify = os.path.join(out_dir, f"주민번호검증_도움컴퍼니_{yymm}.xlsx")
    vout.save(out_verify)

    # ── 4) 일용직 신규 등록 대상 (사원등록 리스트와 이름+생년월일 비교) ──
    out_new = None
    new_workers = []
    if reg_path:
        reg = openpyxl.load_workbook(reg_path, data_only=True).active
        registered = set()
        for r in reg.iter_rows(min_row=2, values_only=True):
            nm = str(r[1] or "").strip()
            if nm:
                registered.add((nm, str(r[3] or "")[:6]))
        seen_p = set()
        for p in daily:
            key = (p["name"], p["resident"][:6])
            if key not in registered and key not in seen_p:
                seen_p.add(key)
                new_workers.append(p)
        nout = openpyxl.Workbook()
        nws = nout.active
        nws.title = "신규등록대상"
        nws.append(["No", "이름", "주민번호", "근무처", "은행", "계좌"])
        for c in nws[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DDDDDD")
        for i, p in enumerate(new_workers):
            nws.append([i + 1, p["name"], p["rrn_disp"], p["work"], p["bank"], p["account"]])
        for col, w in zip("ABCDEF", [5, 12, 18, 18, 10, 20]):
            nws.column_dimensions[col].width = w
        out_new = os.path.join(out_dir, f"일용직_신규등록대상_{yymm}.xlsx")
        nout.save(out_new)

    # ── 리포트 ──
    print(f"[사업소득] {len(biz)}명 / 지급총액 {sum(p['amount'] for p in biz):,.0f} / 소득세 {sum(p['tax'] for p in biz):,.0f} / 지방세 {sum(p['local'] for p in biz):,.0f}")
    print(f"[일용직] {len(daily)}건 / 지급액 {sum(p['amount'] for p in daily):,.0f} / 고용보험 {sum(p['emp'] for p in daily):,.0f}")
    print(f"[주민번호] 문제/확인필요 {len(seen)}건")
    for row in list(seen)[:20]:
        print("  -", row[2], row[3], "→", row[4])
    if reg_path:
        print(f"[일용직 신규등록 필요] {len(new_workers)}명")
        for p in new_workers:
            print("  -", p["name"], p["rrn_disp"])
    print("저장 폴더:", out_dir)
    for f in [out_biz, out_daily, out_verify] + ([out_new] if out_new else []):
        print("  -", os.path.basename(f))


if __name__ == "__main__":
    main()
