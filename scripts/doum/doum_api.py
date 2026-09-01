# 도움컴퍼니 원천세 변환 API 엔진 (웹 마법사용)
# 사용: python doum_api.py <정산원본> <YYYY.MM> <일용직사원등록> <출력폴더>
# stdout으로 JSON 출력. 주민번호 '오류'가 하나라도 있으면 파일을 만들지 않음 (게이트).
import sys
import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
import xlrd
from xlutils.copy import copy as xlcopy
from datetime import datetime, date

TPL_DIR = os.path.dirname(os.path.abspath(__file__))


def clean(v):
    return str(v or "").replace("-", "").strip()


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def floor10(x):
    return int(x // 10) * 10


def fmt_date(v, ym):
    if isinstance(v, datetime):
        return f"{v.year}.{v.month:02d}.{v.day:02d}"
    s = str(v or "")[:10].replace("-", ".")
    return s if len(s) == 10 else ym + ".01"


def check_rrn(rrn, today):
    """(level, message) 목록. level: error|warn"""
    out = []
    d = rrn.replace("-", "").strip()
    if not d.isdigit() or len(d) != 13:
        return [("error", f"자릿수 오류 ({len(d)}자리)")]
    g = d[6]
    cent = {"1": 1900, "2": 1900, "3": 2000, "4": 2000, "5": 1900, "6": 1900, "7": 2000, "8": 2000, "9": 1800, "0": 1800}.get(g)
    if cent is None:
        return [("error", f"성별자리 오류 (7번째={g})")]
    try:
        birth = date(cent + int(d[0:2]), int(d[2:4]), int(d[4:6]))
    except ValueError:
        return [("error", f"생년월일 불가 ({d[0:6]})")]
    if birth > today:
        out.append(("error", "생년월일이 미래 (성별자리 확인)"))
    else:
        age = (today - birth).days // 365
        if age < 14:
            out.append(("error", f"나이 이상 (만 {age}세)"))
        elif age > 90:
            out.append(("warn", f"고령 확인 필요 (만 {age}세)"))
    w = [2, 3, 4, 5, 6, 7, 8, 9, 2, 3, 4, 5]
    s = sum(int(d[i]) * w[i] for i in range(12))
    if (11 - s % 11) % 10 != int(d[12]):
        if g in "1234":
            out.append(("error", "체크섬 불일치 (오타 의심)"))
        else:
            out.append(("warn", "체크섬 불일치 (외국인 신규발급이면 정상 가능)"))
    return out


def main():
    src_path, ym, reg_path, out_dir = sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4]
    year, month = ym.split(".")
    yymm = year[2:] + month
    today = date.today()
    os.makedirs(out_dir, exist_ok=True)

    wb = openpyxl.load_workbook(src_path, data_only=True)
    if "3.3원천세" not in wb.sheetnames or "일용직고용보험" not in wb.sheetnames:
        print(json.dumps({"ok": False, "fatal": "정산 원본에 '3.3원천세'/'일용직고용보험' 시트가 없습니다. 파일을 확인해주세요."}, ensure_ascii=False))
        return

    # ── 파싱 ──
    biz = []
    for r in wb["3.3원천세"].iter_rows(min_row=3, values_only=True):
        name = str(r[2] or "").strip()
        if not name:
            continue
        amount = num(r[9])
        tax = num(r[11]) or (floor10(amount * 0.03) if amount > 0 else 0)
        local = num(r[13]) or (floor10(tax * 0.1) if tax > 0 else 0)
        biz.append({"date": fmt_date(r[16], ym), "name": name, "resident": clean(r[5]),
                    "rrn_disp": str(r[5] or "").strip(), "work": str(r[1] or "").strip(),
                    "amount": amount, "tax": tax, "local": local})

    daily_raw = []
    for r in wb["일용직고용보험"].iter_rows(min_row=3, values_only=True):
        name = str(r[2] or "").strip()
        if not name:
            continue
        d = r[12]
        day = d.day if isinstance(d, datetime) else int(str(d)[8:10] or 1)
        amount = num(r[9])
        emp = num(r[11]) or (floor10(amount * 0.009) if amount > 0 else 0)
        daily_raw.append({"day": day, "name": name, "resident": clean(r[5]),
                          "rrn_disp": str(r[5] or "").strip(), "work": str(r[1] or "").strip(),
                          "bank": str(r[3] or "").strip(), "account": str(r[4] or "").strip(),
                          "amount": amount, "emp": emp})
    merged = {}
    for p in daily_raw:
        key = (p["resident"], p["day"])
        if key in merged:
            merged[key]["amount"] += p["amount"]
            merged[key]["emp"] += p["emp"]
        else:
            merged[key] = dict(p)
    daily = sorted(merged.values(), key=lambda p: (p["day"], p["name"]))

    # ── 1번: 주민번호 검증 (게이트) ──
    people = [("사업소득", p) for p in biz] + [("일용직", p) for p in daily]
    errors, warns = [], []
    seen = set()
    for tag, p in people:
        for level, msg in check_rrn(p["resident"], today):
            row = {"tag": tag, "name": p["name"], "rrn": p["rrn_disp"], "work": p["work"], "msg": msg}
            key = (tag, p["name"], p["resident"], msg)
            if key in seen:
                continue
            seen.add(key)
            (errors if level == "error" else warns).append(row)
    by_rrn = {}
    for tag, p in people:
        by_rrn.setdefault(p["resident"], set()).add(p["name"])
    for rrn, names in by_rrn.items():
        if len(names) > 1:
            errors.append({"tag": "공통", "name": "/".join(sorted(names)), "rrn": rrn, "work": "", "msg": "같은 주민번호에 다른 이름 (오타 의심)"})

    totals = {
        "bizCount": len(biz), "bizAmount": sum(p["amount"] for p in biz),
        "bizTax": sum(p["tax"] for p in biz), "bizLocal": sum(p["local"] for p in biz),
        "dailyCount": len(daily), "dailyAmount": sum(p["amount"] for p in daily),
        "dailyEmp": sum(p["emp"] for p in daily),
    }

    if errors:
        print(json.dumps({"ok": False, "errors": errors, "warns": warns, "totals": totals}, ensure_ascii=False))
        return

    # ── 2번: 일용직 신규 등록 대상 ──
    reg = openpyxl.load_workbook(reg_path, data_only=True).active
    registered = set()
    for r in reg.iter_rows(min_row=2, values_only=True):
        nm = str(r[1] or "").strip()
        if nm:
            registered.add((nm, str(r[3] or "")[:6]))
    new_workers, seen_nw = [], set()
    for p in daily:
        key = (p["name"], p["resident"][:6])
        if key not in registered and key not in seen_nw:
            seen_nw.add(key)
            new_workers.append({"name": p["name"], "rrn": p["rrn_disp"], "work": p["work"], "bank": p["bank"], "account": p["account"]})

    f_new = f"일용직_신규등록대상_{yymm}.xlsx"
    nout = openpyxl.Workbook()
    nws = nout.active
    nws.title = "신규등록대상"
    nws.append(["No", "이름", "주민번호", "근무처", "은행", "계좌"])
    for c in nws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
    for i, p in enumerate(new_workers):
        nws.append([i + 1, p["name"], p["rrn"], p["work"], p["bank"], p["account"]])
    for col, w in zip("ABCDEF", [5, 12, 18, 18, 10, 20]):
        nws.column_dimensions[col].width = w
    nout.save(os.path.join(out_dir, f_new))

    # ── 3번: 사업소득자료입력 ──
    f_biz = f"사업소득자료입력_도움컴퍼니_{yymm}.xls"
    rb = xlrd.open_workbook(os.path.join(TPL_DIR, "사업소득자료입력_template.xls"), formatting_info=True)
    wbo = xlcopy(rb)
    ws = wbo.get_sheet(0)
    for i, p in enumerate(biz):
        r = 3 + i
        ws.write(r, 0, i + 1); ws.write(r, 1, ym); ws.write(r, 2, p["date"])
        ws.write(r, 3, p["name"]); ws.write(r, 4, p["resident"])
        ws.write(r, 7, "기타인적용역자"); ws.write(r, 8, p["date"])
        ws.write(r, 9, p["amount"]); ws.write(r, 10, 3.0)
        ws.write(r, 11, p["tax"]); ws.write(r, 12, p["local"]); ws.write(r, 13, 0.0)
    wbo.save(os.path.join(out_dir, f_biz))

    # ── 4번: 일용직급여자료입력 ──
    f_daily = f"일용직급여자료입력_도움컴퍼니_{yymm}.xls"
    rb2 = xlrd.open_workbook(os.path.join(TPL_DIR, "일용직급여자료입력_template.xls"), formatting_info=True)
    wbo2 = xlcopy(rb2)
    ws2 = wbo2.get_sheet(0)
    for i, p in enumerate(daily):
        r = 4 + i
        ws2.write(r, 0, i + 1); ws2.write(r, 1, p["day"]); ws2.write(r, 2, 1)
        ws2.write(r, 3, p["name"]); ws2.write(r, 4, p["resident"])
        ws2.write(r, 7, 8.0); ws2.write(r, 9, p["amount"]); ws2.write(r, 12, p["emp"])
        ws2.write(r, 16, 0.0); ws2.write(r, 17, 0.0)
    wbo2.save(os.path.join(out_dir, f_daily))

    print(json.dumps({
        "ok": True, "warns": warns, "totals": totals, "newWorkers": new_workers,
        "files": {"newList": f_new, "biz": f_biz, "daily": f_daily},
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
